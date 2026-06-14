from fastapi import APIRouter, UploadFile, File, HTTPException, Depends
from database import q_one, q_all, execute
from psycopg.types.json import Jsonb
from auth import get_current_user
from datetime import datetime, timezone
import openpyxl
import io
import tempfile
import os

router = APIRouter(prefix="/api/port-lineups", tags=["port-lineups"])

HEADER_MARKER = "GEMİ ADI"


def parse_port_report(file_bytes: bytes) -> list:
    """Parse the daily port report Excel file and return structured data."""
    # Save to temp file (openpyxl needs file path or file-like object)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(file_bytes)
        tmp_path = tmp.name

    try:
        wb = openpyxl.load_workbook(tmp_path, data_only=True, read_only=True)
        all_reports = []

        for sheet_name in wb.sheetnames:
            report_date = sheet_name.strip()
            ws = wb[sheet_name]

            ports_data = []
            current_port = None
            current_vessels = []
            expecting_port_name = False

            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
                # Get cell values (columns B through J = indices 1-9)
                cells = {}
                for cell in row:
                    if cell.value is not None and cell.column >= 2 and cell.column <= 10:
                        cells[cell.column] = cell.value

                if not cells:
                    continue

                # Check if this is a header row
                col_b_val = cells.get(2, "")
                if isinstance(col_b_val, str) and col_b_val.strip() == HEADER_MARKER:
                    # Save previous port section
                    if current_port and current_vessels:
                        ports_data.append({
                            "portName": current_port,
                            "vessels": current_vessels
                        })
                    current_port = None
                    current_vessels = []
                    expecting_port_name = True
                    continue

                # If we're expecting a port name (row after header)
                if expecting_port_name:
                    if col_b_val and isinstance(col_b_val, str) and col_b_val.strip():
                        # Check it's not a vessel row (vessel rows have data in col C too)
                        col_c_val = cells.get(3)
                        if not col_c_val:
                            current_port = col_b_val.strip()
                            expecting_port_name = False
                            continue
                        else:
                            # It's actually a vessel row, port name might be missing
                            # Use a default
                            current_port = col_b_val.strip()
                            expecting_port_name = False
                            # Fall through to process as vessel

                # Process vessel data row
                if current_port is not None:
                    vessel_name = cells.get(2, "")
                    loading_port = cells.get(3, "")
                    arrival_date_raw = cells.get(4)
                    status = cells.get(5, "")
                    operation = cells.get(6, "")
                    cargo = cells.get(7, "")
                    bl_tonnage = cells.get(8)
                    buyer = cells.get(9, "")
                    seller = cells.get(10, "")

                    # Must have at least vessel name or loading port to be valid
                    if not vessel_name and not loading_port:
                        continue

                    # Parse arrival date
                    arrival_date_str = ""
                    if isinstance(arrival_date_raw, datetime):
                        arrival_date_str = arrival_date_raw.strftime("%d.%m.%Y")
                    elif isinstance(arrival_date_raw, str) and arrival_date_raw.strip():
                        arrival_date_str = arrival_date_raw.strip()

                    # Parse tonnage
                    tonnage = None
                    if bl_tonnage is not None:
                        try:
                            tonnage = float(bl_tonnage)
                        except (ValueError, TypeError):
                            tonnage = None

                    vessel = {
                        "vesselName": str(vessel_name).strip() if vessel_name else "",
                        "loadingPort": str(loading_port).strip() if loading_port else "",
                        "arrivalDate": arrival_date_str,
                        "status": str(status).strip() if status else "",
                        "operation": str(operation).strip() if operation else "",
                        "cargo": str(cargo).strip() if cargo else "",
                        "blTonnage": tonnage,
                        "buyer": str(buyer).strip() if buyer else "",
                        "seller": str(seller).strip() if seller else "",
                    }
                    current_vessels.append(vessel)

            # Save last port section
            if current_port and current_vessels:
                ports_data.append({
                    "portName": current_port,
                    "vessels": current_vessels
                })

            if ports_data:
                all_reports.append({
                    "reportDate": report_date,
                    "ports": ports_data
                })

        wb.close()
        return all_reports
    finally:
        os.unlink(tmp_path)


@router.post("/upload")
async def upload_port_report(file: UploadFile = File(...), current_user=Depends(get_current_user)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are supported")

    file_bytes = await file.read()
    try:
        reports = parse_port_report(file_bytes)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse Excel file: {str(e)}")

    if not reports:
        raise HTTPException(status_code=400, detail="No valid port data found in the file")

    # Clear existing data and insert new
    execute("DELETE FROM port_lineups")
    now = datetime.now(timezone.utc)
    for report in reports:
        execute(
            "INSERT INTO port_lineups (report_date, ports, uploaded_at, uploaded_by) VALUES (%s, %s, %s, %s)",
            (report["reportDate"], Jsonb(report["ports"]), now, current_user.get("username", "unknown")),
        )

    return {
        "message": f"Successfully uploaded {len(reports)} daily reports",
        "dates": [r["reportDate"] for r in reports],
        "totalPorts": sum(len(r["ports"]) for r in reports),
        "totalVessels": sum(
            sum(len(p["vessels"]) for p in r["ports"])
            for r in reports
        )
    }


@router.get("/dates")
async def get_report_dates(current_user=Depends(get_current_user)):
    dates = [r["report_date"] for r in q_all("SELECT DISTINCT report_date FROM port_lineups")]

    # Sort dates descending (newest first)
    def parse_date(d):
        try:
            return datetime.strptime(d.strip(), "%d.%m.%Y")
        except ValueError:
            return datetime.min

    dates.sort(key=parse_date, reverse=True)
    return {"dates": dates}


@router.get("/report/{report_date}")
async def get_report(report_date: str, current_user=Depends(get_current_user)):
    row = q_one("SELECT report_date, ports, uploaded_at, uploaded_by FROM port_lineups WHERE report_date = %s", (report_date,))
    if not row:
        raise HTTPException(status_code=404, detail=f"No report found for date {report_date}")
    return {"reportDate": row["report_date"], "ports": row["ports"],
            "uploadedAt": row["uploaded_at"].isoformat() if row["uploaded_at"] else None,
            "uploadedBy": row["uploaded_by"]}


@router.get("/summary")
async def get_summary(current_user=Depends(get_current_user)):
    """Get a summary of latest report: port names, vessel counts."""
    # Get the latest report date
    dates = [r["report_date"] for r in q_all("SELECT DISTINCT report_date FROM port_lineups")]
    if not dates:
        return {"latestDate": None, "ports": [], "totalVessels": 0}

    def parse_date(d):
        try:
            return datetime.strptime(d.strip(), "%d.%m.%Y")
        except ValueError:
            return datetime.min

    dates.sort(key=parse_date, reverse=True)
    latest_date = dates[0]

    row = q_one("SELECT ports FROM port_lineups WHERE report_date = %s", (latest_date,))
    if not row:
        return {"latestDate": latest_date, "ports": [], "totalVessels": 0}
    doc = {"ports": row["ports"]}

    port_summary = []
    total_vessels = 0
    for port in doc.get("ports", []):
        vessel_count = len(port.get("vessels", []))
        total_vessels += vessel_count
        port_summary.append({
            "portName": port["portName"],
            "vesselCount": vessel_count
        })

    return {
        "latestDate": latest_date,
        "ports": port_summary,
        "totalVessels": total_vessels,
        "totalDates": len(dates)
    }


# ─── Monthly Line-Up ───

UPLOAD_DIR = os.path.join(os.environ.get("UPLOAD_DIR", "/app/backend/uploads"), "monthly_lineups")
try:
    os.makedirs(UPLOAD_DIR, exist_ok=True)
except OSError:
    pass


def parse_monthly_excel(file_bytes: bytes, filename: str):
    """Parse monthly lineup Excel into port-grouped vessel data (same format as daily)."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(file_bytes)
        tmp_path = tmp.name
    try:
        wb = openpyxl.load_workbook(tmp_path, data_only=True, read_only=True)
        all_vessels = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header_row_idx = None
            col_map = {}
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                vals = [str(c).strip() if c is not None else "" for c in row]
                upper_vals = [v.upper() for v in vals]
                if "GEMİ ADI" in upper_vals or "GEMI ADI" in upper_vals:
                    header_row_idx = i
                    for j, v in enumerate(upper_vals):
                        if v in ("GEMİ ADI", "GEMI ADI"): col_map["vessel"] = j
                        elif v in ("YÜKLEME LİMANI", "YUKLEME LIMANI"): col_map["loadPort"] = j
                        elif v in ("GELİŞ TARİHİ", "GELIS TARIHI"): col_map["arrival"] = j
                        elif v in ("TAHLİYE/YÜKLEME", "TAHLIYE/YUKLEME"): col_map["op"] = j
                        elif v in ("YÜKÜ", "YUKU"): col_map["cargo"] = j
                        elif v in ("B/L TONAJI",): col_map["tonnage"] = j
                        elif v in ("ALICI",): col_map["buyer"] = j
                        elif v in ("SATICI",): col_map["seller"] = j
                        elif v in ("RAPOR TARIHI",): col_map["reportDate"] = j
                        elif v in ("LIMAN",): col_map["port"] = j
                    continue
                if header_row_idx is None:
                    continue
                if not any(vals):
                    continue
                vessel_name = vals[col_map["vessel"]].strip() if "vessel" in col_map and col_map["vessel"] < len(vals) else ""
                if not vessel_name:
                    continue
                def safe_get(key):
                    idx = col_map.get(key)
                    if idx is not None and idx < len(vals):
                        return vals[idx].strip()
                    return ""
                def parse_date_val(val):
                    if not val:
                        return ""
                    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d.%m.%Y"):
                        try:
                            return datetime.strptime(val, fmt).strftime("%d.%m.%Y")
                        except (ValueError, TypeError):
                            continue
                    return val
                tonnage_str = safe_get("tonnage").replace(",", "").replace(".", "").strip()
                try:
                    tonnage = int(float(tonnage_str)) if tonnage_str else 0
                except (ValueError, TypeError):
                    tonnage = 0
                port_name = safe_get("port") or sheet_name.upper()
                all_vessels.append({
                    "portName": port_name.upper(),
                    "vesselName": vessel_name.upper(),
                    "loadingPort": safe_get("loadPort").upper(),
                    "arrivalDate": parse_date_val(safe_get("arrival")),
                    "reportDate": parse_date_val(safe_get("reportDate")),
                    "operation": safe_get("op"),
                    "cargo": safe_get("cargo"),
                    "blTonnage": tonnage,
                    "buyer": safe_get("buyer"),
                    "seller": safe_get("seller"),
                    "status": "",
                })
        wb.close()
        # Group by port
        ports_map = {}
        for v in all_vessels:
            pn = v.pop("portName")
            if pn not in ports_map:
                ports_map[pn] = {"portName": pn, "vessels": []}
            ports_map[pn]["vessels"].append(v)
        ports = sorted(ports_map.values(), key=lambda p: len(p["vessels"]), reverse=True)
        return {"ports": ports, "totalVessels": len(all_vessels), "totalPorts": len(ports)}
    finally:
        os.unlink(tmp_path)


@router.post("/monthly/upload")
async def upload_monthly_lineup(file: UploadFile = File(...), current_user=Depends(get_current_user)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files supported")
    file_bytes = await file.read()
    try:
        parsed = parse_monthly_excel(file_bytes, file.filename)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse: {str(e)}")
    if not parsed or not parsed.get("ports"):
        raise HTTPException(status_code=400, detail="No data found in Excel file")

    # Store file on disk
    import uuid
    doc_id = uuid.uuid4()
    ext = file.filename.rsplit('.', 1)[-1] if '.' in file.filename else 'xlsx'
    stored_name = f"monthly_{doc_id.hex}.{ext}"
    with open(os.path.join(UPLOAD_DIR, stored_name), "wb") as f:
        f.write(file_bytes)

    execute(
        "INSERT INTO monthly_lineups (id, file_name, stored_file_name, ports, total_vessels, total_ports, uploaded_at, uploaded_by) "
        "VALUES (%s, %s, %s, %s, %s, %s, %s, %s)",
        (doc_id, file.filename, stored_name, Jsonb(parsed["ports"]), parsed["totalVessels"], parsed["totalPorts"],
         datetime.now(timezone.utc), current_user.get("username", "unknown")),
    )
    return {
        "id": str(doc_id),
        "fileName": file.filename,
        "totalPorts": parsed["totalPorts"],
        "totalVessels": parsed["totalVessels"],
    }


@router.get("/monthly/list")
async def list_monthly_lineups(current_user=Depends(get_current_user)):
    rows = q_all("SELECT id, file_name, stored_file_name, total_vessels, total_ports, uploaded_at, uploaded_by "
                 "FROM monthly_lineups ORDER BY uploaded_at DESC")
    return [{"id": str(r["id"]), "fileName": r["file_name"], "storedFileName": r["stored_file_name"],
             "totalVessels": r["total_vessels"], "totalPorts": r["total_ports"],
             "uploadedAt": r["uploaded_at"].isoformat() if r["uploaded_at"] else None,
             "uploadedBy": r["uploaded_by"]} for r in rows]


@router.get("/monthly/{doc_id}")
async def get_monthly_lineup(doc_id: str, current_user=Depends(get_current_user)):
    row = q_one("SELECT * FROM monthly_lineups WHERE id = %s", (doc_id,))
    if not row:
        raise HTTPException(status_code=404, detail="Not found")
    doc = {"id": str(row["id"]), "fileName": row["file_name"], "storedFileName": row["stored_file_name"],
           "ports": row["ports"], "totalVessels": row["total_vessels"], "totalPorts": row["total_ports"],
           "uploadedAt": row["uploaded_at"].isoformat() if row["uploaded_at"] else None,
           "uploadedBy": row["uploaded_by"]}
    # If old format with sheets (in data), re-parse from file
    if not doc.get("ports") and (row.get("data") or {}).get("sheets") and row["stored_file_name"]:
        fpath = os.path.join(UPLOAD_DIR, row["stored_file_name"])
        if os.path.exists(fpath):
            with open(fpath, "rb") as f:
                parsed = parse_monthly_excel(f.read(), row["stored_file_name"])
            doc["ports"] = parsed["ports"]
            doc["totalVessels"] = parsed["totalVessels"]
            doc["totalPorts"] = parsed["totalPorts"]
    return doc


@router.delete("/monthly/{doc_id}")
async def delete_monthly_lineup(doc_id: str, current_user=Depends(get_current_user)):
    row = q_one("SELECT stored_file_name FROM monthly_lineups WHERE id = %s", (doc_id,))
    if not row:
        raise HTTPException(status_code=404, detail="Not found")
    stored = row["stored_file_name"]
    if stored:
        fpath = os.path.join(UPLOAD_DIR, stored)
        if os.path.exists(fpath):
            os.remove(fpath)
    execute("DELETE FROM monthly_lineups WHERE id = %s", (doc_id,))
    return {"message": "Deleted"}


@router.get("/monthly/{doc_id}/download")
async def download_monthly_lineup(doc_id: str, current_user=Depends(get_current_user)):
    from fastapi.responses import FileResponse
    row = q_one("SELECT file_name, stored_file_name FROM monthly_lineups WHERE id = %s", (doc_id,))
    if not row or not row["stored_file_name"]:
        raise HTTPException(status_code=404, detail="File not found")
    fpath = os.path.join(UPLOAD_DIR, row["stored_file_name"])
    if not os.path.exists(fpath):
        raise HTTPException(status_code=404, detail="File not on disk")
    return FileResponse(fpath, filename=row["file_name"] or "monthly_lineup.xlsx", media_type="application/octet-stream")
